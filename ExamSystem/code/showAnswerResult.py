#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@author: Python Killer
@license: (C) Copyright 2022-2022
@contact: nanyanzheng@link.cuhk.edu.cn
@File : showAnswerResult.py
@Time : 2022/8/22 22:07
@Software: PyCharm
@Describe: 用户答题正确情况
"""

import pandas as pd
from openpyxl import load_workbook
import logging

# 从excel中取出存储的答案,存在字典中
# 返回值：字典：key是题号，value是answer对象
from code.model.Answer import Answer


def __read_answer():
    df_answer = pd.read_excel('excel_all.xlsx', sheet_name='Answer')
    answer_list = dict()
    for index, row in df_answer.iterrows():
        answer_list[row['problemid']] = Answer(row['problemid'], row['type'], row['answer'])
    return answer_list


# 前端调用的函数：用户选择和答案匹配，并把正确和错误情况写入user表
# @Param：
#   name：用户名称；
#   user_choice: {题号:'用户选择'}用户答题的选择
# @return: 无
def match_answer(name, user_choice):
    logging.info(f'user [{name}] choose: {str(user_choice)}')
    answer_list = __read_answer()
    right = dict()
    wrong = dict()
    for k, v in user_choice.items():
        answer = answer_list[k]
        type = answer.type
        if type == 3:
            v = ''.join(sorted(v))
        if v == answer.answer:
            right[k] = type
        else:
            wrong[k] = type
    # 把right和wrong写表
    __write_user_result(name, right, wrong)


# 读取user表数据
def __read_user():
    df_user = pd.read_excel('excel_all.xlsx', sheet_name='User')
    return df_user


# 把答题正确情况写入user表的right和wrong列
# @Param:
#   nama:用户;
#   right:作答正确的题号;
#   wrong: 作答错误的题号
# @return: 无
def __write_user_result(name, right, wrong):
    # 获取当前user表数据
    df_user = pd.read_excel('excel_all.xlsx', sheet_name='User')
    # 以下是写表准备
    wb = load_workbook(r'excel_all.xlsx')  # 打开指定excel文件
    sheet = wb["User"]  # 指定表
    for index, row in df_user.iterrows():
        if row['name'] == name: # 表中用户是当前作答的用户
            sheet.cell(index + 2, 4, str(right))  # right列写入作答正确的题号
            sheet.cell(index + 2, 5, str(wrong))  # wrong列写入作答错误的题号
            break
    # 保存数据
    wb.save(r"excel_all.xlsx")


if __name__ == '__main__':
    choice = {101:'A',102:'B'}
    match_answer('test_user',choice)
    # __write_user_result('test_user', '2,3,4', 'a,b,c')
